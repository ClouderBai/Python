import psycopg2
import openpyxl
from openpyxl.styles import Font
import os
import uuid
import boto3
import json
import re
import sys

# from model import QueryModel
# os.environ["database"] = "idoctor_compare"
# os.environ["user"] = "postgres"
# os.environ["password"] = "Win2008"
# os.environ["host"] = "192.168.100.46"
# os.environ["port"] = "5432"
# os.environ["aws_access_key_id"]=""
# os.environ["aws_secret_access_key"]=""
# os.environ["region_name"]="cn-northwest-1"
# os.environ["bucket_name"]="lly-cn-ibu-idoctor-qa-public"
s3 = boto3.resource(service_name='s3', aws_access_key_id=os.environ["aws_access_key_id"],
                    aws_secret_access_key=os.environ["aws_secret_access_key"], region_name=os.environ["region_name"])


def upload_single_file_to_s3(s3_key: str, file_path: str):
    print("begin to upload {0},key:{1}", file_path, s3_key)
    bucket_name = os.environ["bucket_name"]
    with open(file_path, 'rb') as f:
        file_stream = f.read()
        try:
            s3.Bucket(bucket_name).put_object(
                Key=s3_key, Body=file_stream)
            print("{} upload done".format(s3_key))
        except Exception as e:
            print("upload {} error:{}".format(s3_key, e))


def export_to_excel(querys=[], folder_path=''):
    print("export_to_excel")
    database = os.environ["database"]
    user = os.environ["user"]
    pwd = os.environ["password"]
    host = os.environ["host"]
    port = os.environ["port"]
    try:
        conn = psycopg2.connect(database=database, user=user, password=pwd, host=host, port=port)
        print("Opened database successfully")
        cur = conn.cursor()
        for query in querys:
            try:
                cur.execute(query["query"])
                data = cur.fetchall()
                wb = openpyxl.Workbook()
                sheet = wb.active
                # col_names = []
                for index in range(len(cur.description)):
                    # col_names.append()
                    sheet.cell(row=1, column=index + 1, value=cur.description[index].name)
                # column_names = [row[0] for row in cur ]
                for rowno, row in enumerate(data, start=1):
                    for colno, cell_value in enumerate(row, start=1):
                        # 处理NULL为NULL
                        if (cell_value == None):
                            sheet.cell(row=rowno + 1, column=colno, value='NULL')
                        else:  # 避免转换时候把空值转为字符串
                            if (isinstance(cell_value, str) and cell_value.lower() == "NULL".lower()):
                                sheet.cell(row=rowno + 1, column=colno,
                                           value='NULLStr')
                            else:  # 正常存储
                                if (type(cell_value) == bytes):
                                    cell_value = "binary"
                                # 处理含有\ 特殊字符
                                # 处理特殊字符并去除添加的引号
                                if (isinstance(cell_value, str)):
                                    cell_value = repr(cell_value)[1:-1]
                                # #bool to str
                                # if(isinstance(cell_value,bool)):
                                #     cell_value=str(cell_value)
                                sheet.cell(row=rowno + 1, column=colno,
                                           value=cell_value)

                        # .value = cell_value
                sheet.row_dimensions[1].font = Font(bold=True)
                if query['name']:
                    sheet.title = query['name']
                # 同名excel文件不让同时读取 即使在不同目录
                file_path = folder_path + "/" + query['name'] + '-pg.xlsx'
                print(file_path)
                create_result = wb.save(file_path)
                print(create_result)
                s3_key = "public/tmp/dbcompare/{}".format(query['name'] + '-pg.xlsx')
                upload_single_file_to_s3(s3_key=s3_key, file_path=file_path)
            except Exception as e1:
                # print("table{} row:{} col:{} failed".format(query["name"],rowno,colno))
                print("table {} failed".format(query["name"]))
                print(e1)
                sys.exit(1)
                # pass

        cur.close()
        conn.close()
    except Exception as e:
        cur.close()
        conn.close()
        print(e)


# def test():
#     with open("./hello_world/test_data.json",'r',encoding="utf-8") as fq:
#         # t=fq.readline()
#         querys=json.load(fq)
#         pathFolder = "C:/WorkSpace/db/data/pgsql"
#         export_to_excel(querys=querys, folder_path=pathFolder)


# test()


def lambda_handler(event, context):
    # print(event)
    # folder_path = "/tmp/{}".format(uuid.uuid4())
    folder_path = "/tmp"
    export_to_excel(event, folder_path=folder_path)
